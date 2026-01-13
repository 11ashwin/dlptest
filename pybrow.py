import os
import platform
import subprocess
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

class BrowserDetector:
    def __init__(self):
        self.system = platform.system()
        self.browsers = []
        
    def get_version_windows(self, exe_path):
        """Get version info for Windows executables"""
        try:
            info = subprocess.check_output(
                f'wmic datafile where name="{exe_path.replace(chr(92), chr(92)*2)}" get Version /value',
                shell=True,
                stderr=subprocess.DEVNULL
            ).decode('utf-8', errors='ignore')
            version = info.split('=')[1].strip() if '=' in info else 'Unknown'
            return version
        except:
            return 'Unknown'
    
    def get_version_unix(self, command):
        """Get version info for Unix-based systems"""
        try:
            result = subprocess.run(
                command,
                shell=True,
                capture_output=True,
                text=True,
                timeout=5
            )
            return result.stdout.strip() or 'Unknown'
        except:
            return 'Unknown'
    
    def detect_windows_browsers(self):
        """Detect browsers on Windows"""
        browser_paths = {
            'Google Chrome': [
                r'C:\Program Files\Google\Chrome\Application\chrome.exe',
                r'C:\Program Files (x86)\Google\Chrome\Application\chrome.exe',
                os.path.expandvars(r'%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe')
            ],
            'Mozilla Firefox': [
                r'C:\Program Files\Mozilla Firefox\firefox.exe',
                r'C:\Program Files (x86)\Mozilla Firefox\firefox.exe'
            ],
            'Microsoft Edge': [
                r'C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe',
                r'C:\Program Files\Microsoft\Edge\Application\msedge.exe'
            ],
            'Opera': [
                os.path.expandvars(r'%LOCALAPPDATA%\Programs\Opera\opera.exe'),
                r'C:\Program Files\Opera\opera.exe'
            ],
            'Brave': [
                os.path.expandvars(r'%LOCALAPPDATA%\BraveSoftware\Brave-Browser\Application\brave.exe'),
                r'C:\Program Files\BraveSoftware\Brave-Browser\Application\brave.exe'
            ],
            'Vivaldi': [
                os.path.expandvars(r'%LOCALAPPDATA%\Vivaldi\Application\vivaldi.exe'),
                r'C:\Program Files\Vivaldi\Application\vivaldi.exe'
            ],
            'Internet Explorer': [
                r'C:\Program Files\Internet Explorer\iexplore.exe',
                r'C:\Program Files (x86)\Internet Explorer\iexplore.exe'
            ]
        }
        
        for browser, paths in browser_paths.items():
            for path in paths:
                if os.path.exists(path):
                    version = self.get_version_windows(path)
                    self.browsers.append({
                        'Browser': browser,
                        'Path': path,
                        'Version': version,
                        'Status': 'Installed'
                    })
                    break
    
    def detect_linux_browsers(self):
        """Detect browsers on Linux"""
        browser_commands = {
            'Google Chrome': ('google-chrome', 'google-chrome --version'),
            'Chromium': ('chromium-browser', 'chromium-browser --version'),
            'Mozilla Firefox': ('firefox', 'firefox --version'),
            'Opera': ('opera', 'opera --version'),
            'Brave': ('brave-browser', 'brave-browser --version'),
            'Vivaldi': ('vivaldi', 'vivaldi --version')
        }
        
        for browser, (cmd, version_cmd) in browser_commands.items():
            try:
                path = subprocess.check_output(
                    f'which {cmd}',
                    shell=True,
                    stderr=subprocess.DEVNULL
                ).decode('utf-8').strip()
                
                if path:
                    version = self.get_version_unix(version_cmd)
                    self.browsers.append({
                        'Browser': browser,
                        'Path': path,
                        'Version': version,
                        'Status': 'Installed'
                    })
            except:
                continue
    
    def detect_mac_browsers(self):
        """Detect browsers on macOS"""
        browser_paths = {
            'Google Chrome': '/Applications/Google Chrome.app',
            'Mozilla Firefox': '/Applications/Firefox.app',
            'Safari': '/Applications/Safari.app',
            'Opera': '/Applications/Opera.app',
            'Brave': '/Applications/Brave Browser.app',
            'Vivaldi': '/Applications/Vivaldi.app',
            'Microsoft Edge': '/Applications/Microsoft Edge.app'
        }
        
        for browser, path in browser_paths.items():
            if os.path.exists(path):
                # Get version from Info.plist
                try:
                    plist_path = os.path.join(path, 'Contents/Info.plist')
                    version = self.get_version_unix(
                        f'defaults read "{plist_path}" CFBundleShortVersionString'
                    )
                except:
                    version = 'Unknown'
                
                self.browsers.append({
                    'Browser': browser,
                    'Path': path,
                    'Version': version,
                    'Status': 'Installed'
                })
    
    def detect_browsers(self):
        """Main detection method"""
        print(f"Detecting browsers on {self.system}...")
        
        if self.system == 'Windows':
            self.detect_windows_browsers()
        elif self.system == 'Linux':
            self.detect_linux_browsers()
        elif self.system == 'Darwin':  # macOS
            self.detect_mac_browsers()
        else:
            print(f"Unsupported operating system: {self.system}")
        
        return self.browsers
    
    def export_to_excel(self, filename='installed_browsers.xlsx'):
        """Export detected browsers to Excel"""
        if not self.browsers:
            print("No browsers detected. Nothing to export.")
            return
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Installed Browsers'
        
        # Add title
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = 'Browser Detection Report'
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Add system info
        ws['A2'] = f'System: {self.system}'
        ws['A3'] = f'Hostname: {platform.node()}'
        ws['A4'] = f'Report Date: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        
        # Headers
        headers = ['Browser', 'Path', 'Version', 'Status']
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=6, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        for row_idx, browser in enumerate(self.browsers, start=7):
            ws.cell(row=row_idx, column=1, value=browser['Browser'])
            ws.cell(row=row_idx, column=2, value=browser['Path'])
            ws.cell(row=row_idx, column=3, value=browser['Version'])
            ws.cell(row=row_idx, column=4, value=browser['Status'])
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        
        # Save workbook
        wb.save(filename)
        print(f"\nReport saved to: {os.path.abspath(filename)}")
        print(f"Total browsers detected: {len(self.browsers)}")

def main():
    detector = BrowserDetector()
    browsers = detector.detect_browsers()
    
    if browsers:
        print(f"\nDetected {len(browsers)} browser(s):")
        for browser in browsers:
            print(f"  - {browser['Browser']} ({browser['Version']})")
    else:
        print("No browsers detected on this system.")
    
    # Export to Excel
    detector.export_to_excel('installed_browsers.xlsx')

if __name__ == '__main__':
    main()
